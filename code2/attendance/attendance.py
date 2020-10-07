#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
attendance.py - last updated 2020-10-04

Create attendance table for a class.

==============================
Copyright 2017-2020 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""

#TODO: Sheet 'Notizen' – At present this sheet will be the second, before
# the individual months. If it should be after all the months, it may be
# necessary to copy it and then delete the original.

### Messages
_BAD_RANGE = ("Ungültige Ferienangabe, Startdatum ist nach dem Enddatum:\n"
        "  {key} = {val}")
_MAKE_ATTENDANCE_TABLE = "Erstelle Anwesenheitstabelle für Klasse {klass}"
_COPY_DATA = "Kopiere Daten von der alten Anwesenheitstabelle"
_UPDATED = "Aktualisierte Anwesenheitstabelle erstellt:  Klasse {klass}"
_NEW_TABLE = "Leere Anwesenheitstabelle erstellt:  Klasse {klass}"
_TABLE_ERROR = ("Code-Fehler in Anwesenheitsvorlage:"
                " {pinfo}.{cell0} ≠ {month}.{cell1}\n  file: {tpfile}")
_BAD_FORMULA = ("Ungültige Formelbezeichnung in Anwesenheitsvorlage, cell:"
        " {month}.{cell}\n   file: {tpfile}")
_WARN_CHANGED_DAY_TAG = ("{day}. {month}: problematische Änderung der"
        " Tagesbezeichnung von '{tag1}' zu '{tag2}'")
_MISSING_PUPIL = "Schüler(in) {pid} ist nicht mehr in der Tabelle"


# Instead of these lists, one could use datetime functions
_months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli",
        "August", "September", "Oktober", "November", "Dezember"]
_tags = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun", "Jul",
        "Aug", "Sep", "Okt", "Nov", "Dez"]


import sys, os, builtins, datetime
if __name__ == '__main__':
    # Enable package import if running as module
    this = sys.path[0]
    sys.path[0] = os.path.dirname(this)

import datetime
from collections import OrderedDict
from copy import copy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle

from local.base_config import SCHOOLYEAR_MONTH_1
import local.attendance_config as ATTENDANCE
from core.base import Dates
from core.pupils import Pupils
from tables.spreadsheet import Spreadsheet


#### Spreadsheet Functions ####
# Note that "," must be used as separator, not ";"!
# For libreoffice: formula0 = '=SUM({sheet1}.{cell}:{sheet2}.{cell})'
formulae = {
    '*f1': '=COUNTIF({col1}{row}:{col2}{row},"{char}")',
    '*f2': '=SUM({col1}{row}:{col2}{row})',
}
# for the total number of school days in a month
formulaS = ('=COUNTIF({col1}{row}:{col2}{row},"+")+COUNTIF({col1}{row}:{col2}{row},"K")'
            '+COUNTIF({col1}{row}:{col2}{row},"P")+COUNTIF({col1}{row}:{col2}{row},"!")')
# for the final summary, cumulative values from all months
formula0 = '=SUM({sheet1}:{sheet2}!{cell})'


class AttendanceError(Exception):
    pass


class Month:
    """Manage information, especially names for the months of the school year.
    The calendar year is adjusted to the month of the school year.
    """
    def __init__ (self, schoolyear, num=None):
        self._firstmonth = SCHOOLYEAR_MONTH_1
        self._month1 = self._boundmonth (num or self._firstmonth)
        self._schoolyear = schoolyear
        self._month = self._month1

    def month (self):
        """Return the index of the month, in the range 1 to 12.
        """
        return self._month

    def __str__ (self):
        return _months[self._month - 1]

    def tag (self):
        return _tags[self._month - 1]

    @staticmethod
    def _boundmonth (m):
        return ((m-1) % 12) + 1

    def year (self):
        return (self._schoolyear if (self._month < self._firstmonth
                    or self._firstmonth == 1)
                else self._schoolyear - 1)

    def increment (self, delta = 1):
        self._month = self._boundmonth(self._month + delta)

    def last_tag (self):
        return _tags[self._boundmonth(self._month1 - 1) - 1]


class Table:
    """openpyxl based spreadsheet handler ('.xlsx'-files).
    """
    def __init__ (self, filepath, sheet=None):
        self._wb = load_workbook (filepath + '.xlsx')
        # If I don't want separate sheet objects, I need some other way
        # of referencing them.
        # Here I just use the names to reference the underlying objects.
        self._sheets = OrderedDict ()
        for sheetObject in self._wb:
            self._sheets [sheetObject.title] = sheetObject


    def _getTable (self, sheet):
        return self._sheets [sheet] if sheet else self._wb.worksheets [0]


    def getSheet (self, ix):
        return self._wb.sheetnames [ix]


    def makeStyle (self, name, copyfromcell, sheet=None):
        ws = self._getTable (sheet)
        newstyle = NamedStyle (name=name)
        newstylecell = ws [copyfromcell]
        newstyle.font = copy (newstylecell.font)
        newstyle.fill = copy (newstylecell.fill)
        newstyle.border = copy (newstylecell.border)
        newstyle.alignment = copy (newstylecell.alignment)
        newstyle.number_format = newstylecell.number_format
        newstyle.protection = copy (newstylecell.protection)
        return newstyle


    def getCell (self, celltag, sheet=None):
        ws = self._getTable (sheet)
        return ws [celltag].value


    def setCell (self, celltag, value=None, style=None, sheet=None):
        ws = self._getTable (sheet)
        cell = ws [celltag]
        cell.value = value
        if style:
            cell.style = style


    def getRowHeight (self, row, sheet=None):
        ws = self._getTable (sheet)
        return ws.row_dimensions [row].height


    def setRowHeight (self, row, height, sheet=None):
        ws = self._getTable (sheet)
        ws.row_dimensions [row].height = height


    def mergeCells (self, crange, sheet=None):
        """The row and column indexes are 1-based,
        """
        ws = self._getTable (sheet)
        #print ("$$$", crange, sheet)
        ws.merge_cells (crange)


    def copy_sheet (self, sheet, name):
        """Make a copy of the named sheet, placing it after the current last sheet.
        """
        ws = self._wb.copy_worksheet (self._getTable (sheet))
        ws.title = name
        self._sheets [name] = ws


    def page_setup (self, sheet, landscape=False, fitHeight=False, fitWidth=False):
        ws = self._getTable (sheet)
        ws.page_setup.orientation = (ws.ORIENTATION_LANDSCAPE if landscape
                else ws.ORIENTATION_PORTRAIT)
        if fitHeight or fitWidth:
            wsprops = ws.sheet_properties
            wsprops.pageSetUpPr = PageSetupProperties (fitToPage=True)
#            ws.page_setup.fitToPage = True
            if not fitHeight:
                ws.page_setup.fitToHeight = False
            elif not fitWidth:
                ws.page_setup.fitToWidth = False


    def save(self, fileinfo):
        self._wb.save(fileinfo)


    def remove_sheet(self, sheet):
#        self._wb.remove_sheet(self._getTable(sheet))
        self._wb.remove(self._getTable(sheet))



class AttendanceTable:
    """Manage the building of an attendance table for the given class.
    """
    @classmethod
    def makeAttendanceTable (cls, schoolyear, klass, oldsheet = None):
        REPORT(_MAKE_ATTENDANCE_TABLE.format(klass = klass))
        at = cls(schoolyear, klass)
        at.setupSheet()
        hols = readHols(schoolyear)
        for i in range(12):
            at.addsheet(hols)
        at.removeMonthTemplate()
        if oldsheet:
            REPORT(_COPY_DATA)
            at.copy_old_sheet(oldsheet)
        newsheet = at.save()
        if oldsheet:
            REPORT(_UPDATED.format(klass = klass))
        else:
            REPORT(_NEW_TABLE.format(klass = klass))
        return newsheet


    def __init__(self, schoolyear, klass):
        self._class = klass
        self._year = schoolyear
        self._month = Month(self._year)
        # Get the template
        self._tpath = os.path.join(RESOURCES, 'templates',
                ATTENDANCE.TEMPLATE)
        self._table = Table(self._tpath)
        self._wsInfo = self._table.getSheet(0)
        self._wsMonth = self._table.getSheet(1)


    def setupSheet(self):
        tag1 = self._month.tag()
        tag2 = self._month.last_tag()
        # First entry row for a pupil, get its height
        self._row0 = int(ATTENDANCE.attendance_row_pupils)
        self._rwH = self._table.getRowHeight(self._row0)
        # Get cell styles from existing cells:
        self._st_id = self._table.makeStyle("s0id",
                ATTENDANCE.attendance_style_id)
        self._st_name = self._table.makeStyle("s0name",
                ATTENDANCE.attendance_style_name)
        self._st_sum = self._table.makeStyle("s0sum",
                ATTENDANCE.attendance_style_sum)
        self._st_N = self._table.makeStyle("s0N",
                ATTENDANCE.attendance_style_N)
        self._st_W = self._table.makeStyle("s0W",
                ATTENDANCE.attendance_style_W)
        self._st_F = self._table.makeStyle("s0F",
                ATTENDANCE.attendance_style_F)
        self._st_X = self._table.makeStyle("s0X",
                ATTENDANCE.attendance_style_X)
        self._st_T = self._table.makeStyle("s0T",
                ATTENDANCE.attendance_style_T)
        # Write date, class and group
        startyear_cell = ATTENDANCE.attendance_cell_date1
        endyear_cell = ATTENDANCE.attendance_cell_date2
        self._table.setCell(endyear_cell, self._year)
        if startyear_cell:
            self._table.setCell(startyear_cell, self._year - 1)
        classText = "Klasse %s" % self._class
        self._table.setCell(ATTENDANCE.attendance_cell_class, classText)
        # Month sheet
        self._daystartcol = ATTENDANCE.attendance_col_daystart
        self._dayendcol = get_column_letter(
                column_index_from_string(self._daystartcol) + 30)
        self._table.setCell(ATTENDANCE.attendance_cell_classM,
                classText, sheet = self._wsMonth)
        # Special formula for summing days:
        daysum_cell = ATTENDANCE.attendance_cell_daysum
        self._validdaysrow = ATTENDANCE.attendance_row_days
        self._table.setCell(daysum_cell, formulaS.format(
                    row = self._validdaysrow,
                    col1 = self._daystartcol, col2 = self._dayendcol),
                sheet = self._wsMonth)
        # Set up total school days on summary sheet:
        self._table.setCell(ATTENDANCE.attendance_cell_totaldays,
                formula0.format(sheet1 = tag1, sheet2 = tag2,
                cell = daysum_cell))
        # Get the start columns for the formula cells
        dcol0 = column_index_from_string(ATTENDANCE.attendance_datacol0)
        dcol1 = column_index_from_string(ATTENDANCE.attendance_datacol1)
        ndcols = int(ATTENDANCE.attendance_datacols)
        # Read formula-tags and formula-ids from second sheet,
        # check tags against first sheet and fetch corresponding formula template:
        fml = []
        fcoderow = ATTENDANCE.attendance_row_codes
        fcodelist = []
        row = self._row0
        for colx in range(ndcols):
            # The "code" letters appear in both sheets, probably in different
            # columns, but I assume in the same row.
            column0 = get_column_letter(dcol0 + colx)
            column1 = get_column_letter(dcol1 + colx)
            celltag0 = column0 + fcoderow
            celltag1 = column1 + fcoderow
            # "Code" letter for this column:
            fcode = self._table.getCell(celltag0)
            fcodelist.append(fcode)
            if self._table.getCell(celltag1, sheet = self._wsMonth) != fcode:
                raise AttendanceError(_TABLE_ERROR.format(
                        tpfile = self._tpath,
                        pinfo = self._wsInfo, cell0 = celltag0,
                        month = self._wsMonth, cell1 = celltag1))
            # Now select the formula
            celltag = column1 + str(row)
            try:
                f = formulae[self._table.getCell(celltag, self._wsMonth)]
            except e:
                raise AttendanceError(_BAD_FORMULA.format(
                        tpfile = self._tpath,
                        month = self._wsMonth, cell = celltag)) from e
            fml.append (f)
        # Add pupil rows, and remember rows
        self._pupilRows = {}
        pupilDataList = Pupils(self._year).classPupils(self._class)
        for pdata in pupilDataList:
            pid = pdata['PID']
            self._pupilRows[pid] = row
            self._table.setRowHeight(row, self._rwH)
            self._table.setRowHeight(row, self._rwH, self._wsMonth)
            n1, n2 = pdata['FIRSTNAME'], pdata['LASTNAME']
#NOTE: The column of these cells is not a configuration item
            self._table.setCell("A%d" % row, pid, self._st_id)
            self._table.setCell("B%d" % row, n1, self._st_name)
            self._table.setCell("C%d" % row, n2, self._st_name)
            self._table.setCell("B%d" % row, n1, self._st_name,
                    sheet = self._wsMonth)
            self._table.setCell("C%d" % row, n2, self._st_name,
                    sheet = self._wsMonth)
            # The formula cells:
            for colx in range(ndcols):
                col = get_column_letter(dcol0 + colx)
                colM = get_column_letter(dcol1 + colx)
                rowstr = str(row)
                # Cell in summary page
                self._table.setCell(col + rowstr, formula0.format(
                        sheet1=tag1, sheet2=tag2,
                        cell=colM + rowstr),
                        self._st_sum)
                # Cell in month page
                f = fml[colx]
                if f:
                    self._table.setCell(colM + rowstr,
                            f.format(row=rowstr,
                                    col1 = self._daystartcol,
                                    col2 = self._dayendcol,
                                    char = fcodelist[colx]),
                            self._st_sum, self._wsMonth)
                colx += 1
            row += 1
        self._rowlimit = row


    def addsheet(self, hols):
        ws = self._month.tag()
        self._table.copy_sheet(self._wsMonth, ws)
        # 1st row
        self._table.setCell(ATTENDANCE.attendance_cell_monthM,
                "%s %04d" % (str(self._month), self._month.year()),
                        sheet = ws)
        col0x = column_index_from_string(self._daystartcol) - 1
        # Build a list of style items for the space to fill,
        # according to the calendar
        year = self._month.year()
        mon = self._month.month()
        stlist = []
        for day in range(33):
            col = get_column_letter(col0x + day)
            try:
                date = datetime.date(year, mon, day)
                if date.weekday() > 4:
                    # Weekend
                    stlist.append(self._st_W)
                elif date in hols:
                    # Holiday weekday
                    stlist.append(self._st_F)
                else:
                    # A normal schoolday
                    stlist.append(self._st_N)
                    self._table.setCell(col + self._validdaysrow, "+",
                            sheet=ws)
            except:
                # Invalid date (or other "error" ...)
                stlist.append(self._st_X)

        # For each pupil, style the attendance cells, and the margin cells
        for row in range(self._row0, self._rowlimit):
            col = col0x
            for st in stlist:
                self._table.setCell(get_column_letter(col) + str(row),
                        style = st, sheet = ws)
                col += 1

        self._table.page_setup(ws, landscape = True, fitHeight = True,
                fitWidth = True)
        self._month.increment()
        return ws


    def save(self, filepath = None):
        if filepath:
            self._table.save(filepath + '.xlsx')
            return None
        else:
            virtual_workbook = BytesIO()
            self._table.save(virtual_workbook)
            return virtual_workbook.getvalue()


    def removeMonthTemplate(self):
        self._table.remove_sheet(self._wsMonth)


    def copy_old_sheet(self, filepath):
        """Transfer data from an older version of an attendance table.
        This would be needed if a pupil is added during a school year.
        <filepath> is the full path to the old file.
        """
        # To avoid too much confusion, all row and column indexes start at 1,
        # unless explicitly stated (use of <Spreadsheet> methods).
        # First get the old tables
        table = Spreadsheet(filepath)

        # Copy pupil attendance data.
        # Need to know the last pupil row in the old file ...
        row = self._row0
        nrows = table.colLen()
        # The month sheets come after the overview and 'notes' sheets:
        months = table.getTableNames()[2:]

        # Copy valid day lines
        col = column_index_from_string(self._daystartcol)
        for month in months:
            ws = table.getTable(month)
            for i in range(31):
                colA = get_column_letter(col + i)
                A1 = colA + self._validdaysrow
                val = table.getABValue(A1, ws)
                newval = self._table.getCell(A1, sheet = month)
                # Compare old and new values, warn if necessary
                if newval:  # ('+')
                    # ok: +, ~, P, K  //  not ok: <empty>, !
                    if (not val) or val == '!':
                        REPORT(_WARN_CHANGED_DAY_TAG.format(
                                day = i, month = month,
                                tag1 = '+', tag2 = val or '<leer>'))
                else:
                    # ok: <empty>, !, P, K  //  not ok: +, ~
                    if val == '+' or val == '~':
                        REPORT(_WARN_CHANGED_DAY_TAG.format(
                                day = i, month = month,
                                tag1 = '<leer>', tag2 = val))
                self._table.setCell(A1, val, sheet = month)

        # Copy pupil attendance data
        while row <= nrows:
            # Get pupil id
            pid = table.getABValue("A%d" % row)
            if pid:
                # <endrow> is used to find the index of the first row
                # after the pupil list
                endrow = row + 1
                for month in months:
                    ws = table.getTable(month)
                    # copy pupil's attendance data for each month
                    for i in range(31):
                        colA = get_column_letter(col + i)
                        val = table.getABValue(colA + str(row), ws)
                        if val:
                            row2 = self._pupilRows.get(pid)
                            if row2:
                                self._table.setCell(colA + str(row2),
                                        val, sheet = month)
                            else:
                                raise AttendanceError(_MISSING_PUPIL.format(
                                        pid = pid))
            row += 1

        # Additional notes
        note_sheet = ATTENDANCE.attendance_sheet_notes
        ws = table.getTable(note_sheet)
        nrows = table.colLen(ws)
        ncols = table.rowLen(ws)
        # Copy whole lines
        for row in range(2, nrows + 1):
            for colx in range(ncols):                  # colx is 0-based
                val = table.getValue(row-1, colx, ws)  # 0-based indexing
                if val:
                    # Write to new table in the corresponding row
                    col = get_column_letter(colx + 1)  # colx is 0-based
                    self._table.setCell(col + str(row), val,
                            sheet = note_sheet)

        return True



def readHols(schoolyear):
    """Return a <set> of <datetime.date> instances for all holidays in the
    calendar file for the given year. The dates are initially in isoformat
    (YYYY-MM-DD).
    """
    deltaday = datetime.timedelta(days = 1)
    hols = set()
    for k, v in Dates.get_calendar(schoolyear).items():
        if k[0] == '_':
            if type(v) == str:
                # single date
                hols.add(datetime.date.fromisoformat(v))
            else:
                d1, d2 = map(datetime.date.fromisoformat, v)
                if d1 >= d2:
                    raise AttendanceError(_BAD_RANGE.format(
                            key = k, val = v))
                while True:
                    hols.add(d1)
                    d1 += deltaday
                    if d1 > d2:
                        break
    return hols



if __name__ == '__main__':
    from core.base import init
    init('TESTDATA')

    _year = 2016
    print ("HOLIDAYS", _year)
    for d in sorted(readHols(_year)):
        print (" ---", d.isoformat())

    print("  ++++++++++++++++++++++++++++++++")
    _klass = '11'
    table = AttendanceTable.makeAttendanceTable(_year, klass = _klass)
    fpath = os.path.join(DATA, 'testing', 'Attendance_%s.xlsx' % _klass)
    with open(fpath, 'wb') as fh:
            fh.write(table)
