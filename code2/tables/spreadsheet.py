### python >= 3.7
# -*- coding: utf-8 -*-
"""
tables/spreadsheet.py

Last updated:  2020-10-04

Spreadsheet file reader, returning all cells as strings.
For reading, simple tsv files (no quoting, no escapes), Excel files (.xlsx)
and Open Document files (.ods) are supported.

Spreadsheet file writer, table contains only strings (and empty cells).
For writing, only simple tsv files (no quoting, no escapes) are supported.

Dates are read and written as strings in the format 'yyyy-mm-dd'.

=+LICENCE=============================
Copyright 2020 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

=-LICENCE========================================
"""

### Messages
_UNSUPPORTED_FILETYPE   = "Nicht unterstützer Dateityp ({ending}):\n   {path}"
_TABLENOTFOUND          = "Tabellendatei existiert nicht:\n   {path}"
_MULTIPLEMATCHINGFILES  = "Mehrere passende Dateien:\n   {path}"
_TABLENOTREADABLE       = "Tabellendatei konnte nicht eingelesen werden:\n   {path}"
_INVALIDSHEETNAME       = "Ungültige Tabellenname: '{name}'"
_INVALIDCELLNAME        = "Ungültiger Zellenbezeichnung: '{name}'"
_INVALID_FILE           = "Ungültige oder fehlerhafte Datei"
_NO_TYPE_EXTENSION      = "Dateityp-Erweiterung fehlt: {fname}"
_DUPLICATECOLUMNNAME    = "Spaltenname doppelt vorhanden: {name}"

import sys, os, datetime, re
if __name__ == '__main__':
    # Enable package import if running as module
    this = sys.path[0]
    sys.path[0] = os.path.dirname(this)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from tables.simple_ods_reader import OdsReader
from tables.dictuple import dictuple


def tsvWriter(dbtable):
    """Write a tab-separated-value table from a list of rows,
    each row is a list of cell values (string only, or <None>).
    Return the table as a <bytes> object (utf-8).
    The elements may not contain tabs or newlines. These will just be
    stripped out.
    """
    def bfilter(text):
        return (re.sub(b'\t\n\r', b'', text.encode('utf-8'))
                if text else b'')

    rowlist = []
    # INFO lines
    for row in dbtable.info:
        rowlist.append(b'\t'.join([bfilter(f) for f in row]))
    # HEADER line
    rowlist.append(b'\t'.join([b'' if f[0] == '$' else bfilter(f)
            for f in dbtable.fieldnames()]))
    # DATA lines
    for row in dbtable:
        rowlist.append(b'\t'.join([bfilter(f) for f in row]))
    return b'\n'.join(rowlist) + b'\n'



class TsvReader(dict):
    def __init__ (self, filepath):
        """Read a tab-separated-value table as a list of rows,
        each row is a list of cell values.
        <filepath> can be the path to a tsv file, but it could be an
        <io.BytesIO> object.
        This format doesn't support multiple sheets, the single table
        is named 'TSV' and the resulting instance has only one key, 'TSV'.
        All values are returned as "stripped" strings, except for empty
        cells, these having the value <None>.
        """
        super().__init__()
        if type(filepath) == str:
            with open(filepath, 'rb') as fbi:
                lines = fbi.read().splitlines()
        else:
            lines = filepath.read().splitlines()
        rows = []
        maxlen = 0
        for row_b in lines:
            print(repr(row_b))
            row = [cell.decode('utf-8').strip() or None
                    for cell in row_b.split(b'\t')]
            l = len(row)
            if l > maxlen:
                maxlen = l
            rows.append(row)
        for row in rows:
            dl = maxlen - len(row)
            if dl:
                row += [None] * dl
        self['TSV'] = rows

    def mergedRanges (self, sheetname):
        """Returns an empty list as tsv doesn't support cell merging.
        """
        return []


class XlsReader(dict):
    def __init__ (self, filepath):
        """Read an Excel spreadsheet as a list of rows,
        each row is a list of cell values.
        All sheets in the file are read, a <dict> of sheets
        (name -> row list) is built.

        This is a read-only utility. Formulae, style, etc. are not retained.
        For formulae the last-calculated value is returned.
        All values are returned as strings.
        """
        super().__init__()
        self._mergedRanges = {}
        # Note that <data_only=True> replaces all formulae by their value,
        # which is probably good for reading, but not for writing!
        wb = load_workbook(filepath, data_only=True)
        for wsname in wb.sheetnames:
            ws = wb[wsname]
            rows = []
            for row in ws.iter_rows():
                values = []
                for cell in row:
                    v = cell.value
                    if type(v) == datetime.datetime:
                        v = v.strftime("%Y-%m-%d")
                    elif type(v) == str:
                        v = v.strip()
                        if v == '':
                             v = None
                    elif v != None:
                        v = str(v)
                    values.append(v)
                rows.append(values)
            self[wsname] = rows
            self._mergedRanges[wsname] = ws.merged_cells.ranges

    def mergedRanges(self, sheetname):
        """Returns a list like ['AK2:AM2', 'H33:AD33', 'I34:J34', 'L34:AI34'].
        """
        return self._mergedRanges[sheetname]


class TableError(Exception):
    pass

class Spreadsheet:
    """This class manages a (read-only) representation of a spreadsheet file.
    The individual table/sheet names are available via the method
    <getTableNames()>.
    The currently selected table can be set using the method
    <setTable('sheetname')>.
    The first table is accessed by default. To access others, an optional
    argument must be passed.
    Row length and column length are available via the methods rowLen()
    and colLen().
    The value of a cell is read using <getValue(row, col)>, where <row>
    and <col> are 0-based indexes.
    All cell values are strings, or <None> if empty.
    """
    _SUPPORTED_TYPES = {'tsv': TsvReader, 'ods': OdsReader, 'xlsx': XlsReader}

    @classmethod
    def supportedType(cls, filename):
        """Check the ending of a file name (or path).
        Return <True> if the type is supported, else <False>.
        """
        fsplit = filename.rsplit('.', 1)
        if len(fsplit) == 2:
            if fsplit[1] in cls._SUPPORTED_TYPES:
                return True
        return False


    def __init__(self, filepath):
        """The filepath can be passed with or without type-extension.
        If no type-extension is given, the folder will be searched for a
        suitable file.
        Alternatively, <filepath> may be an in-memory binary stream
        (io.BytesIO) with attribute 'filename' (so that the
        type-extension can be read).
        """
        self._spreadsheet = None
        self._sheetNames = None
        self._table = None
        self.ixHeaderEnd = None

        if type(filepath) == str:
            # realfile = True
            self.filename = os.path.basename(filepath)
            try:
                ending = self.filename.rsplit('.', 1)[1]
            except IndexError as e:
                ending = None
                # No type-extension provided, test valid possibilities
                fpbase = filepath
                for x in self._SUPPORTED_TYPES:
                    fp = '%s.%s' % (fpbase, x)
                    if os.path.isfile(fp):
                        if ending:
                            raise TableError(_MULTIPLEMATCHINGFILES.format(
                                    path = fpbase)) # from e
                        ending = x
                        filepath = fp
                if not ending:
                    raise TableError(_TABLENOTFOUND.format(
                            path = filepath)) # from e
            else:
                # Check that file exists
                if not os.path.isfile(filepath):
                    raise TableError(_TABLENOTFOUND.format(path=filepath))
            self.filepath = filepath

        else:
            # realfile = False
            try:
                self.filename = filepath.filename
            except:
                raise TableError(_INVALID_FILE)
            try:
                ending = self.filename.rsplit('.', 1)[1]
            except:
                raise TableError(_NO_TYPE_EXTENSION.format(
                        fname=self.filename))
            self.filepath = None

        try:
            handler = self._SUPPORTED_TYPES [ending]
        except:
            raise TableError(_UNSUPPORTED_FILETYPE.format(ending=ending))
        try:
            self._spreadsheet = handler(filepath)
        except:
            raise TableError(_TABLENOTREADABLE.format(
                    path=self.filepath or self.filename))

        self._sheetNames = list(self._spreadsheet)
        # Default sheet is the first:
        self._table = self._spreadsheet[self._sheetNames[0]]

    def rowLen(self, table = None):
        if not table:
            table = self._table
        return len(table[0])

    def colLen(self, table = None):
        if not table:
            table = self._table
        return len(table)

    def getValue(self, rx, cx, table = None):
        if not table:
            table = self._table
        return table[rx][cx]

    def getABValue(self, A1, table = None):
        r, c = self.rowcol(A1)
        return self.getValue(r, c, table)

    def getTableNames(self):
        return self._sheetNames

    def _getTable(self, tablename, failerror = True):
        try:
            #print (self._spreadsheet.keys())
            return self._spreadsheet[tablename]
        except:
            if failerror:
                raise TableError(_INVALIDSHEETNAME.format(name=tablename))
            return None

    def setTable(self, tablename):
        table = self._getTable(tablename)
        if table:
            self._table = table
            return True
        else:
            return False

    def dbTable(self, table = None):
        """Read the table as a database-like table (<DBtable>).
        """
        return DBtable(table or self._table)

#Is this needed?
    def getColumnHeaders(self, rowix, table = None):
        """Return a dict of table headers, header -> column index.
        The row containing the headers is passed as argument.
        """
        self.ixHeaderEnd = None
        headers = {}
        for cellix in range(self.rowLen(table)):
            cellV = self.getValue(rowix, cellix, table)
            if cellV:
                if cellV == '#':
                    continue
                if cellV == '!':
                    self.ixHeaderEnd = cellix
                    break
                headers[cellV] = cellix
        return headers

#Is this needed?
    def getMergedRanges(self, tablename):
        return self._spreadsheet.mergedRanges(tablename)


    @staticmethod
    def rowcol(cellname):
        """Return a tuple (row, column) representing a cell position from
        the given reference in the spreadsheet form, e.g. "B12".
        """
        cell = cellname.upper()
        col = -1
        baseval = ord("A")
        i = 0
        for c in cell:
            v = ord(c)
            if v < baseval:
                break
            i += 1
            col = (col + 1) * 26 + v - baseval
        try:
            assert col >= 0
            return(int (cell[i:]) - 1, col)
        except:
            raise TableError(_INVALIDCELLNAME.format(name=cellname))


    @staticmethod
    def cellname(row, col):
        """Return the name of a cell given its coordinates (0-based):
        """
        return get_column_letter(col+1) + str(row+1)



class DBtable:
    """A database-like table, prepared from a list of rows, each row
    being a list of fields.
    A row with '#' in the first column is regarded as a comment
    and ignored.
    The first row with an entry (except '#') in the first column is
    taken as containing the field names. In this row empty fields
    should be avoided. If there are any empty ones, these will be
    allocated numbered tags beginning with '$'.
    All non-comment rows before this header-row are regarded as possible
    general-information rows. They are available as a list of rows –
    without the (empty) first column – as <self.info>.
    All subsequent non-empty rows are taken as records (unless the
    first column contains '#').
    The table is iterable and indexable (returning the rows).
    """
    @staticmethod
    def empty(row):
        """Test for an empty row.
        """
        for cell in row:
            if cell:
                return False
        return True

    def __iter__(self):
        for row in self.rows:
            yield(row)

    def __getitem__(self, i):
        return self.rows[i]

    def __init__(self, table):
        self.rows = []
        self.info = []
        self.header = None
        for row in table:
            if self.empty(row):
                continue
            c1 = row[0]
            if c1 == '#':
                continue
            if self.header:
                self.rows.append(self.header(row))
            elif c1:
                # The field names
                i = 0   # for automatic tagging of unnamed columns
                cols = []
                for f in row:
                    if f:
                        if f in cols:
                            raise TableError(_DUPLICATECOLUMNNAME.format(
                                    name = f))
                        cols.append(f)
                    else:
                        i += 1
                        cols.append('$%02d' % i)
                self.header = dictuple('DBROW', cols)
            else:
                self.info.append(row[1:])

    def fieldnames(self):
        return self.header.fieldnames()


if __name__ == '__main__':
    from core.base import init
    init('TESTDATA')

    import io
    filepath = os.path.join(DATA, 'testing', 'Test1.tsv')
    fname = os.path.basename(filepath)
    tsv = TsvReader(filepath)
    print("\nROWS:")
    for row in tsv['TSV']:
        print(" :::", row)
    print("\n\nAnd now using a file-like object ...\n")
    with open(filepath, 'rb') as fbi:
        bytefile = fbi.read()
    flo = io.BytesIO(bytefile)
    flo.filename = fname
    tsv = TsvReader(flo)
    print("\nROWS:")
    for row in tsv['TSV']:
        print(" :::", row)

    ss = Spreadsheet(filepath)
    dbt = ss.dbTable()
    print("\nINFO:", dbt.info)
    print("\nFIELDS:", dbt.fieldnames())
    print("\nCONTENT:")
    for row in dbt:
        print(" :::", row)
    print("\n*** 4th row:", dbt[3])
#    print("\n*** 20th row:", dbt[19])

    print("\nRemake tsv:")
    print(tsvWriter(dbt).decode('utf-8'))

    print("\nGRADES 10:")
    ss = Spreadsheet(os.path.join(DATA, 'testing', 'Noten_2', 'Noten_10'))
    dbt = ss.dbTable()
    print("\nINFO:", dbt.info)
    print("\nFIELDS:", dbt.fieldnames())
    print("\nCONTENT:")
    for row in dbt:
        print(" :::", row)

    print("\n\nFAIL: This should be an error ...")
    Spreadsheet(filepath.rsplit('.', 1)[0])
