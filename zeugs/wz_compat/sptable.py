### python >= 3.7
# -*- coding: utf-8 -*-
"""
wz_compat/sptable.py

Last updated:  2019-10-25

Manage importing of teaching data to the main subjects matrix.

=+LICENCE=============================
Copyright 2019 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

=-LICENCE========================================
"""

from collections import namedtuple

from wz_core.configuration import Paths
from wz_table.spreadsheet import Spreadsheet

_NOHEADER = "Keine Bezeichnung für Spalte {col} in Tabelle:\n  {path}"
_UNKNOWN_CLASSES = "Unbekannte Klassen: {cname}"
_BADTNAME = "Unbekannte Lehrkraft: {tname} in {data}"

_TIDX = '~'     # prefix for (assumed) non-teaching staff

def readSPTable (filepath):
    sheet = Spreadsheet (filepath, mustexist=True)
    rows = []

    headers = None
    for rowix in range (sheet.colLen ()):
        # Get the value in the first column
        entry1 = sheet.getValue (rowix, 0)
        if not entry1:
            continue
        if headers == None:
            # Read the column headers from this line
            headers = {}
            for cellix in range (sheet.rowLen ()):
                h = sheet.getValue (rowix, cellix)
                if h:
                    headers [h] = cellix
                else:
                    REPORT.Fail (_NOHEADER, col=cellix+1, path=filepath)
            continue
        Row = namedtuple ('Row', headers)
        ### Read the row data
        rowdata = []
        for col in headers.values ():
            try:
                rowdata.append (sheet.getValue (rowix, col))
            except:
                rowdata.append (None)
        rows.append (Row (*rowdata))
    return rows


def importLessons (schoolyear):
    ### classes
    class2cid = {}
    for row in readSPTable (Paths.getYearPath (schoolyear, 'FILE_SP_CLASSES')):
        #klass = row.CLASS
        #ctag = row.CTAG
        #cid = row.CID
        class2cid [row.CLASS] = row.CID

    ### subjects
    stag2sid = {}
    stag2name = {}
    for row in readSPTable (Paths.getYearPath (schoolyear, 'FILE_SP_SUBJECTS')):
        #sname = row.SNAME
        stag = row.SP
        #sid = row.SID
        if row.SID:
            stag2sid [stag] = row.SID
        stag2name [stag] = row.SNAME

    ### teachers
    tname2tid = {}
    tid2tname = {}
    for row in readSPTable (Paths.getYearPath (schoolyear, 'FILE_SP_TEACHERS')):
        #tname = row.TNAME
        #tid = row.TID
        #tclass = row.TCLASS # space-separated list
        #nz = row.NZ
        tid = (_TIDX + row.TID) if row.NZ else row.TID
        tname2tid [row.TNAME] = tid
        tid2tname [tid] = row.TNAME

    ### lessons
    table = {}
    sidtable = {}
    badsubjects = {}
    badclasses = set ()
    badteachers = {}
    for row in readSPTable (Paths.getYearPath (schoolyear, 'FILE_SP_LESSONS')):
        # There can be a list of teachers
        #tnames = row.TNAMES
        #stag = row.SP
        #klass = row.CLASS
        tnames = [t.strip() for t in row.TNAMES.split (',')]
        stag = row.SP

        # The class can be a list.
        classes = []
        for c in row.CLASS.split (','):
            try:
                klass = class2cid [c.strip()]
            except KeyError:
                badclasses.add (c)
                continue
            classes.append (klass)
        if len (classes) == 0:
            continue

        try:
            sid = stag2sid [stag]
        except KeyError:
            try:
                badsubjects [stag2name [stag]].append (klass)
            except:
                badsubjects [stag2name [stag]] = [klass]
            continue

        for tname in [t.strip() for t in row.TNAMES.split (',')]:
            try:
                tid = tname2tid [tname]
            except KeyError:
                try:
                    badteachers [tname].append ((row.CLASS, stag))
                except:
                    badteachers [tname] = [(row.CLASS, stag)]
                continue
            for klass in classes:
                try:
                    t2s = table [klass]
                except:
                    table [klass] = {tid: {sid}}
                else:
                    try:
                        t2s [tid].add (sid)
                    except:
                        t2s [tid] = {sid}

                try:
                    k2c = sidtable [sid]
                except:
                    sidtable [sid] = {klass: {tid}}
                else:
                    try:
                        k2c [klass].add (tid)
                    except:
                        k2c [klass] = {tid}

    if badclasses:
        REPORT.Warn (_UNKNOWN_CLASSES, cname=badclasses)
    for t, data in badteachers.items ():
        REPORT.Warn (_BADTNAME, tname=t, data=data)

    return table, sidtable, tid2tname, badsubjects


import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
class Table:
    """openpyxl based spreadsheet handler ('.xlsx'-files).
    """
    def __init__ (self, schoolyear):
        l = importLessons (schoolyear)
        self.newdata = l [1]
        self.tid2name = l [2]
        self.badsubjects = l [3]
        self.filepath = Paths.getYearPath (_year, 'FILE_SUBJECTS')
        self._wb = load_workbook (self.filepath + '.xlsx')
        self._ws = self._wb.active
        self.makeStyle ()

    def process (self):
        headers = None
        rows = {}
        rowix = -1
        nonteachers = {}    # {tid -> [(class, sid), ...]}
        conflicts = {}      # {class -> [(sid, oldtidset, newtidset), ...]}
        updates = {}        # {class -> [(sid, newtidset), ...]}
        for row in self._ws.iter_rows ():
            rowix += 1
            values = []
            for cell in row:
                v = cell.value
                if type (v) == datetime.datetime:
                    v = v.strftime ("%Y-%m-%d")
                elif type (v) == str:
                    v = v.strip ()
                    if v == '':
                         v = None
                elif v != None:
                    v = str (v)
                values.append (v)
            if not values [0]:
                continue

            if headers == None:
                headers = {}
                col = 0
                for h in values:
                    if h:
                        headers [h] = col
                    else:
                        REPORT.Fail (_NOHEADER, col=col+1, path=self.filepath)
                    col += 1
                continue

            sid = values [0]
            try:
                data = self.newdata [sid]
            except KeyError:
                continue
            for klass in sorted (data):
                colix = headers [klass]
                tids = set ()
                for tid in data [klass]:
                    if tid [0] == _TIDX:
                        try:
                            nonteachers [tid].append ((klass, sid))
                        except:
                            nonteachers [tid] = [(klass, sid)]
                    else:
                        tids.add (tid)
                current = values [headers [klass]]
                if current:
                    currentset = set ([tid.strip () for tid in current.split (',')])
                    if currentset != tids:
                        # Conflicting updates
                        try:
                            conflicts [klass].append ((sid, currentset, tids))
                        except:
                            conflicts [klass] = [(sid, currentset, tids)]
                else:
                    # New entry
                    try:
                        updates [klass].append ((sid, tids))
                    except:
                        updates [klass] = [(sid, tids)]
                    self.setCell (rowix, colix, ','.join (tids))

        return nonteachers, conflicts, updates


    def makeStyle (self):
        # Style for new entries
        _al = Alignment ()
        _al.horizontal = 'center'
        _al.vertical = 'center'
        self.cellstyle = {
                'font': Font (name = "Arial", size = 11, color="B00000"),
                'number_format': "@",
                'alignment': _al,
                'border': Border (left=Side (style='thin'),
                        right=Side (style='thin'),
                        top=Side (style='thin'),
                        bottom=Side (style='thin'))
            }


    def setCell (self, row, col, val):
        cell = self._ws.cell (row=row+1, column=col+1)
        for k, v in self.cellstyle.items ():
            setattr (cell, k, v)
        if isinstance (val, str) and type (val) != str:
            val = str (val)
        cell.value = val


    def save (self):
        self._wb.save (self.filepath + '.xlsx')




_year = 2020
def test_01 ():
    filepath = Paths.getYearPath (_year, 'FILE_SP_TEACHERS')
    REPORT.Test ("\n  --------------\n%s" % repr (readSPTable (filepath)))
    filepath = Paths.getYearPath (_year, 'FILE_SP_CLASSES')
    REPORT.Test ("\n  --------------\n%s" % repr (readSPTable (filepath)))
    filepath = Paths.getYearPath (_year, 'FILE_SP_SUBJECTS')
    REPORT.Test ("\n  --------------\n%s" % repr (readSPTable (filepath)))
    filepath = Paths.getYearPath (_year, 'FILE_SP_LESSONS')
    REPORT.Test ("\n  --------------\n%s" % repr (readSPTable (filepath)))

def test_02 ():
    table, sidtable, tid2name, badsubjects = importLessons (_year)
    REPORT.Test ("\n  ++++++++++++++++++++++++++++++ \n")
    for klass in sorted (table):
        REPORT.Test ("\n ==== Class %s" % klass)
        data = table [klass]
        for tid in sorted (data):
            REPORT.Test (" $$$ %s: %s" % (tid, data [tid]))

    REPORT.Test ("\n  ++++++++++++++++++++++++++++++ \n")
    for sid in sorted (sidtable):
        REPORT.Test ("\n ==== Subject %s" % sid)
        data = sidtable [sid]
        for klass in sorted (data):
            REPORT.Test (" $$$ %s: %s" % (klass, data [klass]))


    REPORT.Test ("\n &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&\n"
            " NOTE: Subjects taught only in blocks are not included!")

def test_03 ():
    REPORT.Test ("\n  %%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%% \n")
    table = Table (_year)
    nonteachers, conflicts, updates = table.process ()
    REPORT.Test ("\n ++++ TO ADD:")
    for klass in sorted (updates):
        REPORT.Test ("    %s: %s" % (klass, repr (updates [klass])))

    REPORT.Test ("\n ++++ NON-TEACHING STAFF:")
    for tid, data in nonteachers.items ():
        REPORT.Test ("    %s: %s" % (table.tid2name [tid], repr (data)))

    REPORT.Test ("\n ++++ DIFFERENCES:")
    for klass in sorted (conflicts):
        REPORT.Test (" -- Class %s:" % klass)
        for sid, currentset, tids in conflicts [klass]:
            REPORT.Test ("      %s: %s -> %s" % (sid,
                    repr (currentset), repr (tids)))

    REPORT.Test ("\n ++++ UNKNOWN SUBJECTS:")
    for klass in sorted (table.badsubjects):
        REPORT.Test ("    %s: %s" % (klass, repr (table.badsubjects [klass])))

#    table.save ()
