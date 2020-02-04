# python >= 3.7
# -*- coding: utf-8 -*-
"""
wz_grades/gradetable.py - last updated 2020-02-04

Create grade tables for display and grade entry.

==============================
Copyright 2020 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""

#TODO: Need to modify the grade tables to use KlassMatrix.


_FIRSTSIDCOL = 4    # Index (0-based) of first sid column
_UNUSED = '/'      # Subject tag for unused column

# Messages
_MISSING_SUBJECT = "Fachkürzel {sid} fehlt in Notentabellenvorlage:\n  {path}"
_NO_TEMPLATE = "Keine Notentabelle-Vorlage für Klasse/Gruppe {ks} in GRADES.GRADE_TABLE_INFO"
_NO_ITEMPLATE = "Keine Noteneingabe-Vorlage für Klasse/Gruppe {ks} in GRADES.GRADE_TABLE_INFO"


import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle

from wz_core.configuration import Paths
from wz_core.pupils import Pupils, Klass
from wz_core.courses import CourseTables
#TODO: replace Table use by KlassMatrix
from wz_table.matrix import KlassMatrix, Table
from .gradedata import getGradeData


def makeGradeTable(schoolyear, term, klass, title):
    """Make a grade table for the given school-class/group.
    <klass> is a <Klass> instance.
    <term> is a string.
    """
    # Info concerning grade tables:
    gtinfo = CONF.GRADES.GRADE_TABLE_INFO
    # Determine table template
    t = klass.match_map(gtinfo.GRADE_TABLE_TEMPLATE)
    if not t:
        REPORT.Fail(_NO_TEMPLATE, ks=klass)
    template = Paths.getUserPath('FILE_GRADE_TABLE_TEMPLATE').replace('*', t)
    table = Table(template)

    ### Insert general info
    table.setCell('B1', title)
    # "Translation" of info items:
    kmap = CONF.TABLES.COURSE_PUPIL_FIELDNAMES
    info = (
        (kmap['SCHOOLYEAR'], str(schoolyear)),
        (kmap['CLASS'], klass.klass),
        (kmap['TERM'], term)
    )
    i, x = 0, 0
    for row in table.rows:
        i += 1
        if row[0] == '#':
            try:
                k, v = info[x]
            except IndexError:
                REPORT.Fail(_TOO_MANY_INFOLINES, n=len(info), path=template)
            x += 1
            table.setCell('B%d' % i, k)
            table.setCell('C%d' % i, v)
        elif row[0]:
            # The subject key line
            break
    if x < len(info):
        REPORT.Fail(_TOO_FEW_INFOLINES, n=len(info), path=template)
    # <row> is the title row, <i> is the row index of the next row (0-based).

    ### Manage subjects
    courses = CourseTables(schoolyear)
    sid2tlist = courses.classSubjects(klass)
#    print ("???1", list(sid2tlist))
    # Go through the template columns and check if they are needed:
    colmap = {}
    col = _FIRSTSIDCOL
    for k in row[_FIRSTSIDCOL:]:
        if k:
            if k in sid2tlist:
                colmap[k] = col
            elif k == _UNUSED:
                table.hideCol(col, i)
            else:
                # Handle extra _tags
                klassmap = gtinfo.get(k)
                if klassmap:
                    m = klass.match_map(klassmap)
                    if m and term in m.split():
                        colmap[k] = col
                    else:
                        table.hideCol(col, i)
                else:
                    table.hideCol(col, i)
        col += 1
#    print("???COLMAP:", colmap)

    ### Add pupils
    pupils = Pupils(schoolyear)
    for pdata in pupils.classPupils(klass):
        while True:
            i += 1
            try:
                if table.rows[i][0] == 'X':
                    break
            except:
                REPORT.Fail(_TOO_FEW_ROWS, path=template)
        _row = str(i+1)
        pid = pdata['PID']
        table.setCell('A' + _row, pid)
        table.setCell('B' + _row, pdata.name())
        table.setCell('C' + _row, pdata['STREAM'])
        # Add existing grades
        gd = getGradeData(schoolyear, pid, term)
#        print("\n???", pid, gd)
        if gd:
            grades = gd['GRADES']
            if grades:
                for k, v in grades.items():
                    try:
                        col = colmap[k]
                    except KeyError:
#                        print("!!! excess subject:", k)
                        continue
                    if k:
                        if k.startswith('__'):
                            # Calculated entry
                            continue
                        table.setCell(get_column_letter(col+1) + _row, v)
    # Delete excess rows
    table.delEndRows(i + 1)

    ### Save file
    table.protectSheet()
    return table.save()



def stripTable(schoolyear, term, klass, title):
    """Build a basic pupil/subject table for entering grades.
    <klass> is a <Klass> instance.
    <term> is a string.
     """
    # Info concerning grade tables:
    gtinfo = CONF.GRADES.GRADE_TABLE_INFO

    ### Determine table template (output)
    t = klass.match_map(gtinfo.GRADE_INPUT_TEMPLATE)
    if not t:
        REPORT.Fail(_NO_ITEMPLATE, ks=klass)
    template = Paths.getUserPath('FILE_GRADE_TABLE_TEMPLATE').replace('*', t)
    table = Table(template)
    table.setCell('B1', title)

    ### Read input table template (for determining subjects and order)
    # Determine table template (input)
    t = klass.match_map(gtinfo.GRADE_TABLE_TEMPLATE)
    if not t:
        REPORT.Fail(_NO_TEMPLATE, ks=klass)
    template0 = Paths.getUserPath('FILE_GRADE_TABLE_TEMPLATE').replace('*', t)
    table0 = Table(template0)
    i, x = 0, 0
    for row0 in table0.rows:
        i += 1
        if row0[0] and row0[0] != '#':
            # The subject key line
            break
    # <row0> is the title row.

    ### Manage subjects
    courses = CourseTables(schoolyear)
    sid2tlist = courses.classSubjects(klass)
#    print ("???1", list(sid2tlist))
    # Find subject line in new file
    i = 0
    for row in table.rows:
        i += 1
        if row[0]:
            # The subject key line
            break
    # <row> is a list of cell values
    istr = str(i)   # row tag
    # Set klass cell
    table.setCell('A' + istr, row[0].replace('*', klass.klass))
    # Go through the template columns and check if they are needed:
    col = 0
    for sid in row0:
        if sid and sid[0] != '_' and sid in sid2tlist:
            sname = courses.subjectName(sid)
            # Add subject
            while True:
                col += 1
                try:
                    if row[col] == 'X':
                        break
                except:
                    REPORT.Fail(_TOO_FEW_COLUMNS, path=template)
            table.setCell(get_column_letter(col+1) + istr, sname)
    # Delete excess columns
    table.delEndCols(col + 1)

    ### Add pupils
    pupils = Pupils(schoolyear)
    for pdata in pupils.classPupils(klass):
        while True:
            i += 1
            try:
                if table.rows[i][0] == 'X':
                    break
            except:
                REPORT.Fail(_TOO_FEW_ROWS, path=template)
        _row = str(i+1)
        table.setCell('A' + _row, pdata.name())
        table.setCell('B' + _row, pdata['STREAM'])
    # Delete excess rows
    table.delEndRows(i + 1)

    ### Save file
    table.protectSheet()
    return table.save()



##################### Test functions
_testyear = 2016
def test_01():
    _term = '1'
    for ks in '11.RS-HS-_', '11.Gym','12.RS-HS', '12.Gym', '13':
        klass = Klass(ks)
        bytefile = makeGradeTable(_testyear, _term, klass, "Noten: 1. Halbjahr")
        filepath = Paths.getYearPath(_testyear, 'FILE_GRADE_FULL', make=-1,
                term=_term).replace('*', str(klass).replace('.', '-')) + '.xlsx'
        with open(filepath, 'wb') as fh:
            fh.write(bytefile)
        REPORT.Test(" --> %s" % filepath)
        bytefile = stripTable(_testyear, _term, klass, "Noten: 1. Halbjahr")
        filepath = Paths.getYearPath(_testyear, 'FILE_GRADE_INPUT', make=-1,
                term=_term).replace('*', str(klass).replace('.', '-')) + '.xlsx'
        with open(filepath, 'wb') as fh:
            fh.write(bytefile)
        REPORT.Test(" --> %s" % filepath)


def test_02():
    _term = '2'
    for ks in '10', '11.Gym', '11.RS-HS-_', '12.RS-HS', '12.Gym', '13':
        klass = Klass(ks)
        bytefile = makeGradeTable(_testyear, _term, klass, "Noten: 1. Halbjahr")
        filepath = Paths.getYearPath(_testyear, 'FILE_GRADE_FULL', make=-1,
                term=_term).replace('*', str(klass).replace('.', '-')) + '.xlsx'
        with open(filepath, 'wb') as fh:
            fh.write(bytefile)
        REPORT.Test(" --> %s" % filepath)
        bytefile = stripTable(_testyear, _term, klass, "Noten: 1. Halbjahr")
        filepath = Paths.getYearPath(_testyear, 'FILE_GRADE_INPUT', make=-1,
                term=_term).replace('*', str(klass).replace('.', '-')) + '.xlsx'
        with open(filepath, 'wb') as fh:
            fh.write(bytefile)
        REPORT.Test(" --> %s" % filepath)
