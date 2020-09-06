#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
attendance.py - last updated 2020-09-05

Create attendance table for a class.

==============================
Copyright 2017-2019 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""

#TODO: Sheet 'Notizen' – At present this sheet will be the second, before
# the individual months. If it should be after all the months, it may be
# necessary to copy it and then delete the original.

import datetime
import os
from collections import OrderedDict
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle

from wz_core.configuration import Paths, ConfigFile
from wz_core.pupils import Pupils, Klass
from wz_table.spreadsheet import Spreadsheet

#### Spreadsheet Functions ####
# Note that "," must be used as separator, not ";"!
# For libreoffice: formula0 = '=SUM({sheet1}.{cell}:{sheet2}.{cell})'
formulae = {
    '*f1': '=COUNTIF({col1}{row}:{col2}{row},"{char}")',
    '*f2': '=SUM({col1}{row}:{col2}{row})',
}
# for the total number of school days in a month
formulaS = ('=COUNTIF({col1}{row}:{col2}{row},"+")+COUNTIF({col1}{row}:{col2}{row},"K")'
            '+COUNTIF({col1}{row}:{col2}{row},"P")+COUNTIF({col1}{row}:{col2}{row},"!")')
# for the final summary, cumulative values from all months
formula0 = '=SUM({sheet1}:{sheet2}!{cell})'


class Month:
    """Manage information, especially names for the months of the school year.
    The calendar year is adjusted to the month of the school year.
    """
    # Instead of these lists, one could use datetime functions
    _months = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli",
            "August", "September", "Oktober", "November", "Dezember"]
    _tags = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun", "Jul",
            "Aug", "Sep", "Okt", "Nov", "Dez"]

    def __init__ (self, schoolyear, num=None):
        self._firstmonth = CONF.MISC.SCHOOLYEAR_MONTH_1.nat (imax=12, imin=1)
        self._month1 = self._boundmonth (num or self._firstmonth)
        self._schoolyear = schoolyear
        self._month = self._month1

    def month (self):
        """Return the index of the month, in the range 1 to 12.
        """
        return self._month

    def __str__ (self):
        return self._months [self._month - 1]

    def tag (self):
        return self._tags [self._month - 1]

    @staticmethod
    def _boundmonth (m):
        return ((m-1) % 12) + 1

    def year (self):
        return (self._schoolyear if (self._month < self._firstmonth
                    or self._firstmonth == 1)
                else self._schoolyear - 1)

    def increment (self, delta = 1):
        self._month = self._boundmonth (self._month + delta)

    def last_tag (self):
        return self._tags [self._boundmonth (self._month1 - 1) - 1]


class Table:
    """openpyxl based spreadsheet handler ('.xlsx'-files).
    """
    def __init__ (self, filepath, sheet=None):
        self._wb = load_workbook (filepath + '.xlsx')
        # If I don't want separate sheet objects, I need some other way
        # of referencing them.
        # Here I just use the names to reference the underlying objects.
        self._sheets = OrderedDict ()
        for sheetObject in self._wb:
            self._sheets [sheetObject.title] = sheetObject


    def _getTable (self, sheet):
        return self._sheets [sheet] if sheet else self._wb.worksheets [0]


    def getSheet (self, ix):
        return self._wb.sheetnames [ix]


    def makeStyle (self, name, copyfromcell, sheet=None):
        ws = self._getTable (sheet)
        newstyle = NamedStyle (name=name)
        newstylecell = ws [copyfromcell]
        newstyle.font = copy (newstylecell.font)
        newstyle.fill = copy (newstylecell.fill)
        newstyle.border = copy (newstylecell.border)
        newstyle.alignment = copy (newstylecell.alignment)
        newstyle.number_format = newstylecell.number_format
        newstyle.protection = copy (newstylecell.protection)
        return newstyle


    def getCell (self, celltag, sheet=None):
        ws = self._getTable (sheet)
        return ws [celltag].value


    def setCell (self, celltag, value=None, style=None, sheet=None):
        ws = self._getTable (sheet)
        cell = ws [celltag]
        cell.value = value
        if style:
            cell.style = style


    def getRowHeight (self, row, sheet=None):
        ws = self._getTable (sheet)
        return ws.row_dimensions [row].height


    def setRowHeight (self, row, height, sheet=None):
        ws = self._getTable (sheet)
        ws.row_dimensions [row].height = height


    def mergeCells (self, crange, sheet=None):
        """The row and column indexes are 1-based,
        """
        ws = self._getTable (sheet)
        #print ("$$$", crange, sheet)
        ws.merge_cells (crange)


    def copy_sheet (self, sheet, name):
        """Make a copy of the named sheet, placing it after the current last sheet.
        """
        ws = self._wb.copy_worksheet (self._getTable (sheet))
        ws.title = name
        self._sheets [name] = ws


    def page_setup (self, sheet, landscape=False, fitHeight=False, fitWidth=False):
        ws = self._getTable (sheet)
        ws.page_setup.orientation = (ws.ORIENTATION_LANDSCAPE if landscape
                else ws.ORIENTATION_PORTRAIT)
        if fitHeight or fitWidth:
            wsprops = ws.sheet_properties
            wsprops.pageSetUpPr = PageSetupProperties (fitToPage=True)
#            ws.page_setup.fitToPage = True
            if not fitHeight:
                ws.page_setup.fitToHeight = False
            elif not fitWidth:
                ws.page_setup.fitToWidth = False


    def save (self, filepath):
        self._wb.save (filepath + '.xlsx')


    def remove_sheet (self, sheet):
#        self._wb.remove_sheet (self._getTable (sheet))
        self._wb.remove (self._getTable (sheet))



class AttendanceTable:
    """Manage the building of an attendance table for the given class.
    """
    @classmethod
    def makeAttendanceTable (cls, schoolyear, klass, oldsheet=None):
        REPORT.Info ("Erstelle Anwesenheitstabelle für Klasse %s" % klass)
        at = cls (schoolyear, klass)
        at.setupSheet ()
        hols = readHols (schoolyear)
        for i in range (12):
            at.addsheet (hols)
        at.removeMonthTemplate ()
        if oldsheet:
            REPORT.Info ("Kopiere Daten von der alten Datei")
            at.copy_old_sheet (oldsheet)
        at.save ()
        if oldsheet:
            REPORT.Info ("Aktualisierte Anwesenheitstabelle erstellt:  Klasse %s" % klass)
        else:
            REPORT.Info ("Leere Anwesenheitstabelle erstellt:  Klasse %s" % klass)


    def __init__ (self, schoolyear, klass):
        self._class = klass
        self._year = schoolyear
        self._month = Month (self._year)
        # Get the template
        self._tpath = Paths.getUserPath ('TEMPLATE_ATTENDANCE_TABLE')
        self._table = Table (self._tpath)
        self._wsInfo = self._table.getSheet (0)
        self._wsMonth = self._table.getSheet (1)


    def setupSheet (self):
        tag1 = self._month.tag ()
        tag2 = self._month.last_tag ()
        # First entry row for a pupil, get its height
        self._row0 = CONF.ATTENDANCE.attendance_row_pupils.nat ()
        self._rwH = self._table.getRowHeight (self._row0)
        # Get cell styles from existing cells:
        self._st_id = self._table.makeStyle ("s0id", CONF.ATTENDANCE.attendance_style_id)
        self._st_name = self._table.makeStyle ("s0name", CONF.ATTENDANCE.attendance_style_name)
        self._st_sum = self._table.makeStyle ("s0sum", CONF.ATTENDANCE.attendance_style_sum)
        self._st_N = self._table.makeStyle ("s0N", CONF.ATTENDANCE.attendance_style_N)
        self._st_W = self._table.makeStyle ("s0W", CONF.ATTENDANCE.attendance_style_W)
        self._st_F = self._table.makeStyle ("s0F", CONF.ATTENDANCE.attendance_style_F)
        self._st_X = self._table.makeStyle ("s0X", CONF.ATTENDANCE.attendance_style_X)
        self._st_T = self._table.makeStyle ("s0T", CONF.ATTENDANCE.attendance_style_T)
        # Write date, class and group
        startyear_cell = CONF.ATTENDANCE.attendance_cell_date1
        endyear_cell = CONF.ATTENDANCE.attendance_cell_date2
        self._table.setCell (endyear_cell, self._year)
        if startyear_cell:
            self._table.setCell (startyear_cell, self._year - 1)
        classText = "Klasse %s" % self._class
        self._table.setCell (CONF.ATTENDANCE.attendance_cell_class, classText)
        # Month sheet
        self._daystartcol = CONF.ATTENDANCE.attendance_col_daystart
        self._dayendcol = get_column_letter (column_index_from_string (self._daystartcol) + 30)
        self._table.setCell (CONF.ATTENDANCE.attendance_cell_classM, classText, sheet = self._wsMonth)
        # Special formula for summing days:
        daysum_cell = CONF.ATTENDANCE.attendance_cell_daysum
        self._validdaysrow = CONF.ATTENDANCE.attendance_row_days
        self._table.setCell (daysum_cell, formulaS.format (
                    row = self._validdaysrow,
                    col1 = self._daystartcol, col2 = self._dayendcol),
                sheet = self._wsMonth)
        # Set up total school days on summary sheet:
        self._table.setCell (CONF.ATTENDANCE.attendance_cell_totaldays,
                formula0.format (sheet1=tag1, sheet2=tag2, cell=daysum_cell))
        # Get the start columns for the formula cells
        dcol0 = column_index_from_string (CONF.ATTENDANCE.attendance_datacol0)
        dcol1 = column_index_from_string (CONF.ATTENDANCE.attendance_datacol1)
        ndcols = CONF.ATTENDANCE.attendance_datacols.nat ()
        # Read formula-tags and formula-ids from second sheet,
        # check tags against first sheet and fetch corresponding formula template:
        fml = []
        fcoderow = CONF.ATTENDANCE.attendance_row_codes
        fcodelist = []
        row = self._row0
        for colx in range (ndcols):
            # The "code" letters appear in both sheets, probably in different
            # columns, but I assume in the same row.
            column0 = get_column_letter (dcol0 + colx)
            column1 = get_column_letter (dcol1 + colx)
            celltag0 = column0 + fcoderow
            celltag1 = column1 + fcoderow
            # "Code" letter for this column:
            fcode = self._table.getCell (celltag0)
            fcodelist.append (fcode)
            if self._table.getCell (celltag1, sheet = self._wsMonth) != fcode:
                REPORT.Error ("Code-Fehler in Anwesenheitsvorlage %s:\n   %s.%s ≠ %s.%s"
                        % (self._tpath, self._wsInfo, celltag0, self._wsMonth, celltag1))
            # Now select the formula
            celltag = column1 + str (row)
            try:
                f = formulae [self._table.getCell (celltag, self._wsMonth)]
            except:
                REPORT.Error ("Ungültige Formelbezeichnung in Anwesenheitsvorlage %s:\n   %s.%s"
                        % (self._tpath, self._wsMonth, celltag))
                f = None
            fml.append (f)
        # Add pupil rows, and remember rows
        self._pupilRows = {}
        pupilDataList = Pupils (self._year).classPupils (Klass(self._class))
        for pdata in pupilDataList:
            pid = pdata ['PID']
            self._pupilRows [pid] = row
            self._table.setRowHeight (row, self._rwH)
            self._table.setRowHeight (row, self._rwH, self._wsMonth)
            n1, n2 = pdata ['FIRSTNAME'], pdata['LASTNAME']
#NOTE: The column of these cells is not a configuration item
            self._table.setCell ("A%d" % row, pid, self._st_id)
            self._table.setCell ("B%d" % row, n1, self._st_name)
            self._table.setCell ("C%d" % row, n2, self._st_name)
            self._table.setCell ("B%d" % row, n1, self._st_name, sheet = self._wsMonth)
            self._table.setCell ("C%d" % row, n2, self._st_name, sheet = self._wsMonth)
            # The formula cells:
            for colx in range (ndcols):
                col = get_column_letter (dcol0 + colx)
                colM = get_column_letter (dcol1 + colx)
                rowstr = str(row)
                # Cell in summary page
                self._table.setCell (col + rowstr, formula0.format (sheet1=tag1, sheet2=tag2,
                        cell=colM + rowstr),
                        self._st_sum)
                # Cell in month page
                f = fml [colx]
                if f:
                    self._table.setCell (colM + rowstr, f.format (row=rowstr,
                                col1 = self._daystartcol, col2 = self._dayendcol,
                                char = fcodelist [colx]),
                            self._st_sum, self._wsMonth)
                colx += 1
            row += 1
        self._rowlimit = row


    def addsheet (self, hols):
        ws = self._month.tag ()
        self._table.copy_sheet (self._wsMonth, ws)
        # 1st row
        self._table.setCell (CONF.ATTENDANCE.attendance_cell_monthM,
                "%s %04d" % (str (self._month), self._month.year ()), sheet=ws)
        col0x = column_index_from_string (self._daystartcol) - 1
        # Build a list of style items for the space to fill, according to the calendar
        year = self._month.year ()
        mon = self._month.month ()
        stlist = []
        for day in range (33):
            col = get_column_letter (col0x + day)
            try:
                date = datetime.date (year, mon, day)
                if date.weekday () > 4:
                    # Weekend
                    stlist.append (self._st_W)
                elif date in hols:
                    # Holiday weekday
                    stlist.append (self._st_F)
                else:
                    # A normal schoolday
                    stlist.append (self._st_N)
                    self._table.setCell (col + self._validdaysrow, "+" , sheet=ws)
            except:
                # Invalid date (or other "error" ...)
                stlist.append (self._st_X)

        # For each pupil, style the attendance cells, and the margin cells
        for row in range (self._row0, self._rowlimit):
            col = col0x
            for st in stlist:
                self._table.setCell (get_column_letter (col) + str (row),
                        style = st, sheet = ws)
                col += 1

        self._table.page_setup (ws, landscape=True, fitHeight=True, fitWidth=True)
        self._month.increment ()
        return ws


    def save (self):
        filepath = Paths.getYearPath (self._year, 'FILE_ATTENDANCE_TABLE',
                make=-1, klass=self._class)
        self._table.save (filepath)


    def removeMonthTemplate (self):
        self._table.remove_sheet (self._wsMonth)


    def copy_old_sheet (self, filepath):
        """Transfer data from an older version of an attendance table.
        This would be needed if a pupil is added during a school year.
        <filepath> is the full path to the old file.
        """
        # To avoid too much confusion, all row and column indexes start at 1,
        # unless explicitly stated (use of <Spreadsheet> methods).
        # First get the old tables
        table = Spreadsheet (filepath)

        # Copy pupil attendance data.
        # Need to know the last pupil row in the old file ...
        row = self._row0
        nrows = table.colLen ()
        # The month sheets come after the overview and 'notes' sheets:
        months = table.getTableNames () [2:]

        # Copy valid day lines
        col = column_index_from_string (self._daystartcol)
        for month in months:
            ws = table.getTable (month)
            for i in range (31):
                colA = get_column_letter (col + i)
                A1 = colA + self._validdaysrow
                val = table.getABValue (A1, ws)
                newval = self._table.getCell (A1, sheet = month)
                # Compare old and new values, warn if necessary
                if newval:  # ('+')
                    # ok: +, ~, P, K  //  not ok: <empty>, !
                    if (not val) or val == '!':
                        REPORT.Warn ("%d. %s: Problematische Änderung der Tagesbezeichnung von '+' zu '%s'"
                                % (i, month, val or '<leer>'))
                else:
                    # ok: <empty>, !, P, K  //  not ok: +, ~
                    if val == '+' or val == '~':
                        REPORT.Warn ("%d. %s: Problematische Änderung der Tagesbezeichnung von '<leer>' zu '%s'"
                                % (i, month, val))
                self._table.setCell (A1, val, sheet = month)

        # Copy pupil attendance data
        while row <= nrows:
            # Get pupil id
            pid = table.getABValue ("A%d" % row)
            if pid:
                # <endrow> is used to find the index of the first row
                # after the pupil list
                endrow = row + 1
                for month in months:
                    ws = table.getTable (month)
                    # copy pupil's attendance data for each month
                    for i in range (31):
                        colA = get_column_letter (col + i)
                        val = table.getABValue (colA + str (row), ws)
                        if val:
                            row2 = self._pupilRows.get (pid)
                            if row2:
                                self._table.setCell (colA + str (row2), val, sheet = month)
                            else:
                                REPORT.Error (("Schüler(in) %s ist nicht"
                                        " mehr in der Tabelle") % pid)
            row += 1

        # Additional notes
        note_sheet = CONF.ATTENDANCE.attendance_sheet_notes
        ws = table.getTable (note_sheet)
        nrows = table.colLen (ws)
        ncols = table.rowLen (ws)
        # Copy whole lines
        for row in range (2, nrows + 1):
            for colx in range (ncols):                  # colx is 0-based
                val = table.getValue (row-1, colx, ws)  # 0-based indexing
                if val:
                    # Write to new table in the corresponding row
                    col = get_column_letter (colx + 1)  # colx is 0-based
                    self._table.setCell (col + str (row), val,
                            sheet = note_sheet)

        return True




def readHols (schoolyear):
    """Return a <set> of <datetime.date> instances for all valid dates in the
    holidays file (configuration item "HOLIDAYS"). The dates are in isoformat
    (YYYY-MM-DD), but also MM-DD is acceptable, in which case the year will be
    added automatically (from the current school year).
    """
    deltaday = datetime.timedelta (days=1)
    kalinfo = ConfigFile (Paths.getYearPath (schoolyear, 'FILE_HOLIDAYS'))
    hols = set ()
    for d0 in kalinfo.SINGLE_DAYS.csplit ('|'):
        d = getDate (schoolyear, d0, dateformat=False)
        if d:
            hols.add (d)
        else:
            raise RuntimeError ("Date Error")

    for r0 in kalinfo.RANGES.csplit ('|'):
        try:
            d01, d02 = kalinfo [r0].csplit ('|')
        except:
            REPORT.Fail ("Ungültige Ferienzeit: %s" % r0)
        d1 = getDate (schoolyear, d01, dateformat=False)
        d2 = getDate (schoolyear, d02, dateformat=False)
        if (not d1) or (not d2):
            raise RuntimeError ("Date Error")
        if d1 >= d2:
            REPORT.Fail (("Ungültige Ferienangabe: %s"
                    "Startdatum ist nach dem Enddatum") % item)
            continue
        while True:
            hols.add (d1)
            d1 += deltaday
            if d1 > d2:
                break

    return hols



def getDate (schoolyear, date0, dateformat=True):
    """Input is either 'm-d' or 'y-m-d'. If no year is given, get it
    from the current school year.
    Also <datetime.datetime> instances are accepted.
    Check that the date is within the currently selected school year.
    If <dateformat> is <False>, return a <datetime.date> instance.
    If <dateformat> is <True>, return the date in 'iso' format (YYYY-MM-DD).
    In case of failure, return <None>.
    """
    m1 = CONF.MISC.SCHOOLYEAR_MONTH_1.nat (imax=12, imin=1)
    y = schoolyear if m1 == 1 else schoolyear - 1
    dstart = datetime.date (y, m1, 1)
    dend = datetime.date (y + 1, m1, 1)
    if type (date0) == str:
        ymd = [int (i) for i in date0.split ("-")]
        if len (ymd) == 2:
            ymd.insert (0, schoolyear - 1 if ymd [0] >= m1 else schoolyear)
        try:
            date = datetime.date (*ymd)
        except:
            REPORT.Error ("Ungültiges Datum: %s" % date0)
            return None
    else:
        # Assume a <datetime.datetime> instance
        date = date0.date ()
    if date < dstart or date >= dend:
        REPORT.Error ("Datum nicht im aktuellen Schuljahr: %s"
                % date0)
        return None
    return date.isoformat () if dateformat else date


##################### Test functions
def test_01 ():
    year = 2021
    print ("Hols", year)
    l = list (readHols (year))
    l.sort ()
    for d in l:
        print (" ---", d.isoformat ())

def test_02 ():
    year = 2021
    AttendanceTable.makeAttendanceTable (year, klass='11')
