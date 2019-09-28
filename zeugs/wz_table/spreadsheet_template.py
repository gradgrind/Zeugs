#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
wz_table/spreadsheet_template.py

Last updated:  2019-06-30

Spreadsheet file generator using a template.
Only Excel files (.xlsx) are supported.

All input and output values are strings,
Dates are handled as strings in the format 'yyyy-mm-dd'.

=+LICENCE=============================
Copyright 2019 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

=-LICENCE========================================
"""

#NOTE: It is possible to use openpyxl without properly 'installing' it.
# Just place the package folder ('openpyxl') in the 'app' folder.
# openpyxl requires jdcal (just the file 'jdcal.py') and et-xmlfile (the
# package folder 'et_xmlfile'), which must then also be included
# in the 'app' folder. et-xmlfile is only needed if lxml is not available.


### Messages
_BADDATE        = ("Ungültiges Datum (JJJJ-MM-TT: {val}) für Feld {key} in:\n"
                    "  Tabelle: {sheet}, Datei: {fpath}")
_KEYMISSING     = ("Kein Wert für Feld {key} in:\n"
                    "  Tabelle: {sheet}, Datei: {fpath}")
_BADREF         = ("Ungültiger Verweis ({ref}) für Feld {key} in:\n"
                    "  Tabelle: {sheet}, Datei: {fpath}")


import os, datetime
#from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def XLS_template (outfile, filepath, fields, sheetname=None):
    """Replace referenced fields in an Excel spreadsheet by the supplied values.

    The input file <filepath> is the template, the path to the resulting
    file is <outfile>. If the containing folder doesn't exist, it will be
    created. If <outfile> exists already, it will be silently overwritten.

    The template will normally be a multisheet spreadsheet. All sheets
    will be protected, so fields which are to be manually edited should
    be tagged as unprotected.

    One of the sheets lists the fields to be accessed externally.
    The text in the first column is the field name (key), the second
    column contains a reference to the field in question as a formula:
    '=sheet-name!cell-name' (e.g. '=Eingabe!E14'). Also cell references
    without sheet-name are possible: '=cell-name'. In that case the sheet
    will be the same one as that from which these entries are read.
    The name of the sheet containing these references is passed in as the
    parameter <sheetname>. If not supplied, the first sheet is assumed.

    The values to be placed in the listed fields are supplied in the
    mapping <fields>: {key -> value}. Both key and value should be strings.
    If a field is missing, the cell will be blanked.
    If a key ends with '_D' the value should be a date, formatted as
    'yyyy-mm-dd'.

    A set of unused field names is returned.
    """
    # Note that <data_only=True> replaces all formulae by their value,
    # which is often good for reading, but not for writing!
    wb = load_workbook (filepath, data_only=False)
    if not sheetname:
        sheetname = wb.sheetnames [0]
    ws = wb [sheetname]
#    print ("?1", sheetname, ws.title)
    fset = set (fields)
    for row in ws.iter_rows ():
        key = row [0].value
        if key:
            if key.startswith ('---'):
                break
            val = row [1].value
#            print ("++", key, val)
            try:
                fset.remove (key)
            except:
                REPORT.Error (_KEYMISSING, key=key,
                        fpath=filepath, sheet=sheetname)
                continue
            else:
                newval = fields [key]
                try:
                    if key.endswith ('_D'):
                        newval = datetime.date (*[int (v)
                                for v in newval.split ('-')])
                except:
                    REPORT.Error (_BADDATE, key=key, val=newval,
                        fpath=filepath, sheet=sheetname)
                    continue
            ref = val.lstrip ('=').split ('!')
            if len (ref) == 1:
                sheet, cadr = sheetname, ref [0]
            elif len (ref) == 2:
                sheet, cadr = ref
            else:
                REPORT.Error (_BADREF, key=key, ref=val,
                        fpath=filepath, sheet=sheetname)
                continue

            wb [sheet] [cadr].value = newval

    for ws in wb:
        ws.protection.enable ()
    folder = os.path.dirname (outfile)
    if not os.path.isdir (folder):
        os.makedirs (folder)
    wb.save (outfile)
    return fset

