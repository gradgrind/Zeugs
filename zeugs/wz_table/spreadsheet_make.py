#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
wz_table/spreadsheet_make.py

Last updated:  2019-10-14

Create a new spreadsheet (.xlsx).

=+LICENCE=============================
Copyright 2017-2019 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

=-LICENCE========================================
"""

import os, datetime
from collections import namedtuple

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import (NamedStyle, PatternFill, Alignment,
        Protection, Font, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import (WorksheetProperties,
        PageSetupProperties)


class NewSpreadsheet:
    FORMAT_DATE = 'DD.MM.YYYY'

    def __init__ (self, sheetName=None):
        # Create the workbook and worksheet we'll be working with
        self._wb = Workbook ()
        self._ws = self._wb.active
        if sheetName:
            self._ws.title = sheetName
        self._unlocked = None   # cache for a <Protection> instance


    @staticmethod
    def cellName (row, col):
        if row == None:
            r = '*r*'
        else:
            r = str (row+1)
        if col == None:
            c = '*c*'
        else:
            c = get_column_letter (col+1)
        return c + r


    @staticmethod
    def completeCellNames (rcstring, row=None, col=None):
        if col != None:
            rcstring = rcstring.replace ('*c*', get_column_letter (col+1))
        if row != None:
            rcstring = rcstring.replace ('*r*', str (row+1))
        return rcstring


    def makeStyle (self, style):
        """Return the attributes of this style in the form needed for
        applying it to a cell. The result is cached in the style object.
        """
        cellstyle = style.cellStyle
        if cellstyle == None:
            cellstyle = {}
            cstyle = style.attributes
            # Font
            try:
                fontname = cstyle ['font']
                f = True
            except KeyError:
                fontname = 'Arial'
                f = False
            try:
                fontsize = int (cstyle ['size'])
                f = True
            except KeyError:
                fontsize = 12
            try:
                fontbold = bool (cstyle ['bold'])
                f = True
            except KeyError:
                fontbold = False
            try:
                fontital = bool (cstyle ['emph'])
                f = True
            except KeyError:
                fontital = False
            try:
                fontcol = cstyle ['fg']
                f = True
            except KeyError:
                fontcol = '000000'
                pass
            if f:
                cellstyle ['font'] = Font (name = fontname,
                        size = fontsize, bold = fontbold,
                        italic = fontital, color=fontcol)

            # "Number format"
            try:
                cellstyle ['number_format'] = cstyle ['number_format']
            except KeyError:
                pass

            # Alignment
            try:
                align = cstyle ['align']
                if align in 'bmt':
                    # Vertical
                    h = 'c'
                    v = align
                    rotate = 90
                else:
                    h = align
                    v = 'm'
                    rotate = None
                cellstyle ['alignment'] = self.alignment (h=h, v=v,
                        rotate=rotate)
            except KeyError:
                pass

            # Border
            try:
                border = cstyle ['border']
                if border == 2:
                    cellstyle ['border'] = self.border (left=0, right=0,
                            top=0, bottom=2)
                elif border == 1:
                    cellstyle ['border'] = self.border ()
            except KeyError:
                pass

            # Background
            try:
                cellstyle ['fill'] = self.background (cstyle ['background'])
            except KeyError:
                pass

            # Validation is not really a style ...
            try:
                valid = cstyle ['valid']
                if valid:
                    # The default is 'locked' so only if <valid> is present
                    # is an action necessary.
                    if not self._unlocked:
                        self._unlocked = Protection (locked=False)
                    # Remove cell protection
                    cellstyle ['protection'] = self._unlocked

                    if type (valid) == list:
                        style.validation = self.dataValidation (valid)
            except KeyError:
                pass

            style.cellStyle = cellstyle
        return cellstyle


    def setCell (self, row, col, val, style=None, isDate=False):
        """Set the cell at the given coordinates to the given value.
        The coordinates start at 0.
        Style objects can be passed as additional arguments.
        """
        cell = self._ws.cell (row=row+1, column=col+1)
        if style:
            cellstyle = self.makeStyle (style)
            for k, v in cellstyle.items ():
                setattr (cell, k, v)

            if style.validation:
                style.validation.add (cell)

        if val != None:
            if isDate:
                # Set cell number format
                cell.number_format = self.FORMAT_DATE
                # Convert to <datetime.date> instance
                cell.value = datetime.date (*[int (v) for v in val.split ('-')])
            else:
# Workaround for probable bug in openpyxl:
                if isinstance (val, str) and type (val) != str:
                    val = str (val)
                cell.value = val


    def setWidth (self, col, width):
        """Set a column width in mm – probably very roughly.
        """
        # The internal width parameter is related to the width of the
        # 'Normal style font'! The conversion factor tries to compensate.
        self._ws.column_dimensions [get_column_letter (col+1)].width = width * 0.5


    def setHeight (self, row, height):
        """Set a row height in mm – probably very roughly.
        """
        # The internal height parameter is related to the height of the
        # 'Normal style font'! The conversion factor tries to compensate.
        self._ws.row_dimensions [row+1].height = height * 2.8


    def merge (self, row0, col0, height, width):
        self._ws.merge_cells (start_row=row0 + 1, start_column=col0 + 1,
                end_row=row0 + height, end_column=col0 + width)


    def dataValidation (self, valList, allow_blank=True):
        """Create a data-validation object with list validation.
        """
        def newValidationList ():
            dv = DataValidation (type='list',
                    formula1 = '"' + ','.join (valList) + '"',
                    allow_blank = allow_blank)

            # Optionally set a custom error message
            #dv.error ='Your entry is not in the list'
            #dv.errorTitle = 'Invalid Entry'

            # Optionally set a custom prompt message
            #dv.prompt = 'Please select from the list'
            #dv.promptTitle = 'List Selection'

            # Add the data-validation object to the worksheet
            self._ws.add_data_validation (dv)
            return dv

        key = tuple (valList) + (allow_blank,)
        try:
            return self._vcache [key]
        except AttributeError:
            # No cache yet
            self._vcache = {}
        except KeyError:
            # No existing validation instance for this key
            pass
        dv = newValidationList ()
        self._vcache [key] = dv
        return dv


    def dataValidationLength (self, chars):
        """Create a data-validation object for a string with maximum
        length validation (chars >= 0) or exact length validation
        (-chars for chars < 0).
        """
        if chars < 0:
            op = 'equal'
            chars = - chars
        else:
            op = 'lessThanOrEqual'
        dv = DataValidation(type='textLength', operator=op, formula1=chars)
        # Optionally set a custom error message
        dv.error ='Entry is too long'

        # Add the data-validation object to the worksheet
        self._ws.add_data_validation (dv)
        return dv


    @staticmethod
    def background (colour):
        return PatternFill (patternType='solid', fgColor=colour)


    @staticmethod
    def alignment (h=None, v=None, rotate=None, indent=None, wrap=None):
        al = Alignment ()
        if h:
            hal = {'l': 'left', 'r': 'right', 'c': 'center'}.get (h)
            if hal:
                al.horizontal = hal
        if v:
            val = {'t': 'top', 'b': 'bottom', 'm': 'center'}.get (v)
            if val:
                al.vertical = val
        if rotate:
            try:
                ral = int (rotate)
                if ral >=0 and ral <= 180:
                    al.textRotation = ral
            except:
                pass
        if indent != None:
            al.indent = float (indent)
        if wrap != None:
            al.wrapText = wrap
        return al


    @staticmethod
    def border (left=1, right=1, top=1, bottom=1):
        """Simple borders. Only supports definition of the sides and thickness.
        The value must lie in the range 0 – 3.
        """
        bstyle = [None, 'thin', 'medium', 'thick']
        return Border (
                left=Side (style=bstyle [left]),
                right=Side (style=bstyle [right]),
                top=Side (style=bstyle [top]),
                bottom=Side (style=bstyle [bottom]))


    def protectSheet (self, pw=None):
        if pw:
            self._ws.protection.set_password (pw)
        else:
            self._ws.protection.enable ()


    def sheetProperties (self, paper='A4', landscape=False,
                        fitWidth=False, fitHeight=False):
        if landscape:
            self._ws.page_setup.orientation = self._ws.ORIENTATION_LANDSCAPE
        self._ws.page_setup.paperSize = getattr (self._ws, 'PAPERSIZE_' + paper)

        # Property settings
        if fitWidth or fitHeight:
            wsprops = self._ws.sheet_properties
            wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True,
                    autoPageBreaks=False)

#            self._ws.page_setup.fitToPage = True
            if not fitWidth:
                self._ws.page_setup.fitToWidth = False
            if not fitHeight:
                self._ws.page_setup.fitToHeight = False


    def freeze (self, row, col):
        self._ws.freeze_panes = self.cellName (row, col)


    def save (self, filepath):
        """Write the spreadsheet to a file.
        The ending '.xlsx' is added automatically if it is not present
        already.
        Return the full filepath.
        """
        fdir = os.path.dirname (filepath)
        fname = os.path.basename (filepath).rsplit ('.', 1) [0] + '.xlsx'
        fp = os.path.join (fdir, fname)
        self._wb.save (fp)
        return fp



class TableStyle:
    def __init__ (self, base=None, **kargs):
        """
        <base> is an existing style (<TableStyle> instance).
        The following kargs are processed:
        <font> is the font name.
        <size> is the font size.
        <bold> (bool) and <emph> (bool) are font styles.
        <fg> is the font colour (RRGGBB).
        <align> is the horizontal (l, c or r) OR vertical (b, m, t) alignment.
            Vertical alignment is for rotated text.
        <background> is a colour in the form 'RRGGBB', default none.
        <border>: Only three border types are supported here:
            0: none
            1: all sides
            2: (thicker) underline
        <number_format>: By default force all cells to text format.
        <valid>: <True> just unlocks cell (removes protection).
            Otherwise it can be a list of valid strings (which will also
            unlock the cell).
#TODO: other types of validation?
        """
        if base == None:
            self.attributes = {}
            # Set default values
            if 'border' not in kargs: kargs ['border'] = 1
            if 'number_format' not in kargs: kargs ['number_format'] = '@'
            if 'align' not in kargs: kargs ['align'] = 'c'
        else:
            self.attributes = base.attributes.copy ()
        self.attributes.update (kargs)

        # These are for the sheet style info (cache),
        # see <NewSpreadsheet.makeStyle>.
        self.cellStyle = None
        self.validation = None
