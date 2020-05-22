### python >= 3.7
# -*- coding: utf-8 -*-
"""
wz_table/spreadsheet.py

Last updated:  2020-05-21

Spreadsheet file reader, returning all cells as strings.

Currently Excel files (.xlsx) and Open Document files (.ods) are supported.

Dates are returned as strings in the format 'yyyy-mm-dd'.

=+LICENCE=============================
Copyright 2017-2020 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.

=-LICENCE========================================
"""

#NOTE: It is possible to use openpyxl without properly 'installing' it.
# Just place the package folder ('openpyxl') in the 'app' folder.
# openpyxl requires jdcal (just the file 'jdcal.py') and et-xmlfile (the
# package folder 'et_xmlfile'), which must then also be included
# in the 'app' folder. et-xmlfile is only needed if lxml is not available.

### Messages
_UNSUPPORTEDFILETYPE    = "Nicht unterstützer Dateityp ({ending}):\n   {path}"
_TABLENOTFOUND          = "Tabellendatei existiert nicht:\n   {path}"
_MULTIPLEMATCHINGFILES  = "Mehrere passende Dateien:\n   {path}"
_TABLENOTREADABLE       = "Tabellendatei konnte nicht eingelesen werden:\n   {path}"
_INVALIDSHEETNAME       = "Ungültige Tabellenname: '{name}'"
_INVALIDCELLNAME        = "Ungültiger Zellenbezeichnung: '{name}'"
_INVALID_FILE           = "Ungültige oder fehlerhafte Datei"
_NO_TYPE_EXTENSION      = "Dateityp-Erweiterung fehlt: {fname}"


import os, datetime
from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .simpleods import OdsReader


class XLS_spreadsheet:
    def __init__ (self, filepath):
        """Read an Excel spreadsheet as a list of rows,
        each row is a list of cell values.
        All sheets in the file are read, an 'OrderedDict' of sheets
        (name -> row list) is built.

        This is a read-only utility. Formulae, style, etc. are not retained.
        For formulae the last-calculated value is returned.
        All values are returned as strings.
        """
        sheets = OrderedDict ()
        self._mergedRanges = OrderedDict ()
        # Note that <data_only=True> replaces all formulae by their value,
        # which is probably good for reading, but not for writing!
        wb = load_workbook (filepath, data_only=True)
        for wsname in wb.sheetnames:
            ws = wb [wsname]
            rows = []
            for row in ws.iter_rows ():
                values = []
                for cell in row:
                    v = cell.value
                    if type (v) == datetime.datetime:
                        v = v.strftime ("%Y-%m-%d")
                    elif type (v) == str:
                        v = v.strip ()
                        if v == '':
                             v = None
                    elif v != None:
                        v = str (v)
                    values.append (v)
                rows.append (values)
            sheets[wsname] = rows
            self._mergedRanges [wsname] = ws.merged_cells.ranges
        self.sheets = sheets

    def mergedRanges (self, sheetname):
        """Returns a list like ['AK2:AM2', 'H33:AD33', 'I34:J34', 'L34:AI34'].
        """
        return self._mergedRanges [sheetname]


class ODS_spreadsheet:
    def __init__ (self, filepath):
        """Read an ".ods" (LibreOffice) spreadsheet as a list of rows,
        each row is a list of cell values.
        All sheets in the file are read, an 'OrderedDict' of sheets
        (name -> row list) is built.

        This is a read-only utility. Formulae, style, etc. are not retained.
        For formulae the last-calculated value is returned.
        All values are returned as strings.
        Numbers which can be represented as integers (xxxx.0) are returned
        as integers (in string form).
        """
        # Read the file to a list form (see <OdsReader>)
        sheets = OdsReader.readOdsFile (filepath)
        # Extract the needed data in the form specified above
        self.sheets = OrderedDict ()
        self._mergedRanges = OrderedDict ()
        for sheetname, rows, mergelist in sheets:
            rowlist = []
            for row in rows:
                cells = []
                for cell in row:
                    if cell == None:
                        cells.append (None)
                    else:
                        ctype, cval, ctext, cformula = cell
                        if ctype == None:
                            v = None
                        elif ctype == 'string':
                            v = ctext.strip ()
                            if v == '':
                                v = None
                        elif ctype == 'float':
                            # Fix for integers returned as floats
                            i = int (cval)
                            v = str (i if i == cval else cval)
                        else:
                            v = str (cval)
                        cells.append (v)
                rowlist.append (cells)
            self.sheets [sheetname] = rowlist
            self._mergedRanges [sheetname] = mergelist


    def mergedRanges (self, sheetname):
        """Returns a list like ['AK2:AM2', 'H33:AD33', 'I34:J34', 'L34:AI34'].
        """
        mrlist = []
        for mr in self._mergedRanges [sheetname]:
            # (rcount, ccount, rs or 1, cs or 1) -> "A2:B4" (for example)
            c1 = get_column_letter (mr [1] + 1)
            r1 = mr [0] + 1
            c2 = get_column_letter (mr [1] + mr [3])
            r2 = mr [0] + mr [2]
            mrlist.append ("%s%d:%s%d" % (c1, r1, c2, r2))
        return mrlist



class Spreadsheet:
    """This class manages a (read-only) representation of a spreadsheet file.
    The individual table/sheet names are available via the method getTableNames().
    The currently selected table can be set using the method setTable('sheetname').
    The first table is accessed primarily. To access others, an optional argument
    must be passed.
    Row length and column length are available via the methods rowLen() and colLen().
    The value of a cell is read using <getValue(row, col)>, where <row> and <col> are
    0-based indexes.
    All cell values are strings, or <None> if empty.
    """
    _SUPPORTED_TYPES = {'ods': ODS_spreadsheet, 'xlsx': XLS_spreadsheet}

    @classmethod
    def supportedType (cls, filename):
        """Check the ending of a file name (or path).
        Return <True> if the type is supported, else <False>.
        """
        fsplit = filename.rsplit ('.', 1)
        if len (fsplit) == 2:
            if fsplit [1] in cls._SUPPORTED_TYPES:
                return True
        return False


    def __init__ (self, filepath, mustexist=True):
        """The filepath can be passed with or without type-extension.
        If no type-extension is given, the folder will be searched for a
        suitable file.
        Alternatively, <filepath> may be a file object with attribute
        'filename' (so that the type-extension can be read).
        """
        self._spreadsheet = None
        self._sheetNames = None
        self._table = None
        self.ixHeaderEnd = None

        if type(filepath) == str:
            # realfile = True
            fname = os.path.basename(filepath)
            try:
                ending = fname.rsplit('.', 1)[1]
            except:
                ending = None
                # No type-extension provided, test valid possibilities
                fpbase = filepath
                for e in self._SUPPORTED_TYPES:
                    fp = '%s.%s' % (fpbase, e)
                    if os.path.isfile (fp):
                        if ending:
                            REPORT.Fail (_MULTIPLEMATCHINGFILES, path=fpbase)
                        ending = e
                        filepath = fp
                if not ending:
                    if mustexist:
                        REPORT.Fail(_TABLENOTFOUND, path=filepath)
                    else:
                        # To handle this, this exception must be caught ...
                        raise FileNotFoundError
                    return
            else:
                # Check that file exists
                if not os.path.isfile (filepath):
                    if mustexist:
                        REPORT.Fail(_TABLENOTFOUND, path=filepath)
                    else:
                        # To handle this, this exception must be caught ...
                        raise FileNotFoundError
                    return
            self.filepath = filepath

        else:
            # realfile = False
            try:
                fname = filepath.filename
            except:
                REPORT.Fail(_INVALID_FILE)
            try:
                ending = fname.rsplit('.', 1)[1]
            except:
                REPORT.Fail(_NO_TYPE_EXTENSION, fname=fname)
            self.filepath = None

        try:
            handler = self._SUPPORTED_TYPES [ending]
        except:
            REPORT.Fail (_UNSUPPORTED_FILETYPE, ending=ending)
        try:
            self._spreadsheet = handler(filepath)
        except:
            raise
            # Error: couldn't read file
            REPORT.Fail (_TABLENOTREADABLE, path=self.filepath or fname)

        self._sheetNames = list (self._spreadsheet.sheets)
        # Default sheet is the first:
        self._table = self._spreadsheet.sheets [self._sheetNames [0]]

    def rowLen (self, table = None):
        if not table:
            table = self._table
        return len (table [0])

    def colLen (self, table = None):
        if not table:
            table = self._table
        return len (table)

    def getValue (self, rx, cx, table = None):
        if not table:
            table = self._table
        return table [rx] [cx]

    def getABValue (self, A1, table = None):
        r, c = self.rowcol (A1)
        return self.getValue (r, c, table)

    def getTableNames (self):
        return self._sheetNames

    def _getTable (self, tablename, failerror=True):
        try:
            #print (self._spreadsheet.sheets.keys ())
            return self._spreadsheet.sheets [tablename]
        except:
            if failerror:
                REPORT.Fail (_INVALIDSHEETNAME, name=tablename)
                assert False
            return None

    def setTable (self, tablename):
        table = self._getTable (tablename)
        if table:
            self._table = table
            return True
        else:
            return False

    def getColumnHeaders (self, rowix, table = None):
        """Return a dict of table headers, header -> column index.
        The row containing the headers is passed as argument.
        """
        self.ixHeaderEnd = None
        headers = {}
        for cellix in range (self.rowLen(table)):
            cellV = self.getValue (rowix, cellix, table)
            if cellV:
                if cellV == '#':
                    continue
                if cellV == '!':
                    self.ixHeaderEnd = cellix
                    break
                headers[cellV] = cellix
        return headers


    def getMergedRanges (self, tablename):
        return self._spreadsheet.mergedRanges (tablename)


    @staticmethod
    def rowcol (cellname):
        """Return a tuple (row, column) representing a cell position from
        the given reference in the spreadsheet form, e.g. "B12".
        """
        cell = cellname.upper ()
        col = -1
        baseval = ord ("A")
        i = 0
        for c in cell:
            v = ord (c)
            if v < baseval:
                break
            i += 1
            col = (col + 1) * 26 + v - baseval
        try:
            assert col >= 0
            return (int (cell[i:]) - 1, col)
        except:
            REPORT.Fail (_INVALIDCELLNAME, name=cellname)
            assert False


    @staticmethod
    def cellname (row, col):
        """Return the name of a cell given its coordinates (0-based):
        """
        return get_column_letter (col+1) + str (row+1)
