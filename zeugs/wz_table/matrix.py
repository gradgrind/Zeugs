# python >= 3.7
# -*- coding: utf-8 -*-
"""
wz_table/matrix.py - last updated 2020-02-04

Edit a table template (xlsx).

The base class is <Table>.

<KlassMatrix> handles pupil-subject matrices.

==============================
Copyright 2020 Michael Towers

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

       http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""

# Messages
_TOO_MANY_INFOLINES = "Zu viele Infozeilen ('#, ...'), {n} erforderlich, in:\n  {path}"
_TOO_FEW_INFOLINES = "Zu wenige Infozeilen ('#, ...'), {n} erforderlich, in:\n  {path}"
_TOO_FEW_COLUMNS = "Noteneingabe-Vorlage hat zu wenige Spalten:\n  {path}"
_TOO_FEW_ROWS = "Noteneingabe-Vorlage hat zu wenige Zeilen:\n  {path}"


import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side, PatternFill, NamedStyle


class Table:
    """openpyxl based spreadsheet handler ('.xlsx'-files).
    """
    @staticmethod
    def columnLetter(i):
        """Return the column letter(s) for the given (0-based) index.
        """
        return get_column_letter(i+1)


    def __init__(self, filepath):
        self.template = filepath + '.xlsx'
        self._wb = load_workbook(self.template)
        self.rows = []
        for row in self._wb.active.iter_rows():
            values = []
            for cell in row:
                v = cell.value
                if isinstance(v, datetime.datetime):
                    v = v.strftime("%Y-%m-%d")
                elif isinstance(v, str):
                    v = v.strip()
                    if v == '':
                         v = None
                elif v != None:
                    v = str(v)
                values.append(v)
            self.rows.append(values)


    def getCell(self, celltag):
        return self._wb.active[celltag].value


    def read(self, row, col):
        """Read the cell at the given position (0-based indexes).
        """
        coltag = self.columnLetter(col)
        return self.getCell(coltag + str(row+1))


    def setCell(self, celltag, value=None, style=None):
        cell = self._wb.active[celltag]
        cell.value = value
        if style:
            cell.style = style


    def write(self, row, col, val):
        """Write to the cell at the given position (0-based indexes).
        """
        coltag = self.columnLetter(col)
        self.setCell(coltag + str(row+1), val)


    def hideCol(self, index, clearheader=None):
        """Hide the given column (0-indexed).
        <clearheader> is the row number (1-based) of a cell to be cleared.
        """
        letter = get_column_letter(index+1)
#TODO: disabled pending hidden column fix ...
#        self._wb.active.column_dimensions[letter].hidden = True
        if clearheader:
            # Clear any existing "subject"
            self.setCell(letter + str(clearheader))
            self.setCell(letter + str(clearheader+1))


    def delEndCols(self, col0):
        """Delete last columns, starting at index <col0> (0-based).
        """
        ndel = len(self.rows[0]) - col0
        if ndel > 0:
            self._wb.active.delete_cols(col0+1, ndel)


    def delEndRows(self, row0):
        """Delete last rows, starting at index <row0> (0-based).
        """
        ndel = len(self.rows) - row0
        if ndel > 0:
            self._wb.active.delete_rows(row0+1, ndel)


    def protectSheet (self, pw=None):
        if pw:
            self._wb.active.protection.set_password(pw)
        else:
            self._wb.active.protection.enable()


    def save (self, filepath=None):
        if filepath:
            self._wb.save(filepath + '.xlsx')
            return None
        else:
            virtual_workbook = BytesIO()
            self._wb.save(virtual_workbook)
            return virtual_workbook.getvalue()



class KlassMatrix(Table):
    """An extension of the <Table> class to deal with pupil-subject tables.
    """

    def setTitle(self, title):
        """The title cell is at a fixed position, "B1". "A1" is empty.
        """
        self.setCell('B1', title)


    def setInfo(self, info):
        i, x = 0, 0
        for row in self.rows:
            i += 1
            if row[0] == '#':
                try:
                    k, v = info[x]
                except IndexError:
                    REPORT.Fail(_TOO_MANY_INFOLINES, n = len(info),
                            path = self.template)
                x += 1
                self.setCell('B%d' % i, k)
                self.setCell('C%d' % i, v)
            elif row[0]:
                # The subject key line
                break
        if x < len(info):
            REPORT.Fail(_TOO_FEW_INFOLINES, n = len(info),
                    path = self.template)
        # <row> is the header row
        self.headers = row
        # <i> is the row index of the next row (0-based),
        # initially immediately after the headers
        self.rowindex = i
        # column index for header column iteration
        self.hcol = 0


    def nextcol(self):
        """Iterate over header columns with 'X' in template.
        """
        while True:
            self.hcol += 1
            try:
                if self.headers[self.hcol] == 'X':
                    return self.hcol
            except:
                REPORT.Fail(_TOO_FEW_COLUMNS, path=self.template)


    def nextrow(self):
        """Iterate over pupil rows ('X' in first column).
        """
        while True:
            self.rowindex += 1
            try:
                if self.rows[self.rowindex][0] == 'X':
                    return self.rowindex
            except:
                REPORT.Fail(_TOO_FEW_ROWS, path=self.template)
